# -*- coding: utf-8 -*-
"""Did an edit change anything it was not asked to change?

Run by `npm run test:xlsx`, after the Rust tests have written before/after pairs
into src-tauri/target/fidelity/.

## Why this is not a Rust test

The Rust side can assert that untouched zip entries come out byte-identical,
and it does. What it cannot honestly check is the FORMATTING of the cells in
the sheet it rewrote: it would be comparing its own output against its own
notion of what a border is, and a wrong notion satisfies both halves. That is
exactly how the border loss went unnoticed for as long as it did.

So a second implementation reads the files. openpyxl has no idea what this
project believes, which is the entire point.

## What is checked

For every cell in every sheet of every pair:

  * the value is unchanged, except in the cell the edit named
  * font, fill, border, alignment and number format are unchanged, everywhere
  * no cell appeared or disappeared

Plus package integrity: every relationship Target must resolve to a part that
exists. A dangling Target is what makes Excel say it repaired the file.
"""
import os
import posixpath
import re
import sys
import zipfile

try:
    from openpyxl import load_workbook
except ImportError:
    print("skip: openpyxl is not installed (pip install openpyxl)")
    sys.exit(0)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PAIRS = os.path.join(ROOT, 'src-tauri', 'target', 'fidelity')

# The cell the Rust harness edits, per fixture. Everything else must be
# untouched. rich/A2 is a STRING, which the writer stores as an inline string
# rather than in the shared table — this is where that gets read back by
# something that is not us.
EDITED = {
    'styled': ('明細', 'B2'),
    'rich': ('明細', 'A2'),
    # A STYLE edit. The value does not change; the format of that one cell does,
    # and of nothing else. A format is a shared index, so "restyle A2" turning
    # into "restyle every cell that looked like A2" is the failure to catch.
    'restyle': ('明細', 'A2'),
}

# Fixtures where the edit changes the format rather than the value.
RESTYLED = {'restyle'}


def fmt(c):
    f, fl, b, a = c.font, c.fill, c.border, c.alignment
    return (
        ('bold', f.bold), ('italic', f.italic), ('size', f.size), ('name', f.name),
        ('color', getattr(f.color, 'rgb', None) if f.color else None),
        ('fill', getattr(fl.fgColor, 'rgb', None) if fl and fl.fgColor else None),
        ('pattern', fl.patternType if fl else None),
        ('numfmt', c.number_format),
        ('halign', a.horizontal), ('valign', a.vertical), ('wrap', a.wrap_text),
        ('border', tuple((getattr(b, side).style if getattr(b, side) else None)
                         for side in ('left', 'right', 'top', 'bottom'))),
    )


def cells(path):
    wb = load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                out[(ws.title, c.coordinate)] = (fmt(c), c.value)
    return out


def dangling_relationships(path):
    z = zipfile.ZipFile(path)
    names = set(z.namelist())
    bad = []
    for n in sorted(names):
        if not n.endswith('.rels'):
            continue
        base = posixpath.dirname(posixpath.dirname(n))
        xml = z.read(n).decode('utf8', 'replace')
        for rel in re.findall(r'<Relationship\b[^>]*>', xml):
            if 'TargetMode="External"' in rel:
                continue
            m = re.search(r'Target="([^"]+)"', rel)
            if not m:
                continue
            t = m.group(1)
            if t.startswith(('http', 'mailto', 'file:', '#')):
                continue
            resolved = (t.lstrip('/') if t.startswith('/')
                        else posixpath.normpath(posixpath.join(base, t)))
            if resolved not in names:
                bad.append('%s -> %s' % (n, t))
    return bad


def check(stem):
    before_p = os.path.join(PAIRS, stem + '-before.xlsx')
    after_p = os.path.join(PAIRS, stem + '-after.xlsx')
    a, b = cells(before_p), cells(after_p)
    edited = EDITED[stem]
    problems = []

    gone = sorted(set(a) - set(b))
    new = sorted(set(b) - set(a))
    if gone:
        problems.append('セルが消えた: %s' % gone[:10])
    if new:
        problems.append('セルが増えた: %s' % new[:10])

    restyling = stem in RESTYLED
    for k in sorted(set(a) & set(b)):
        fa, va = a[k]
        fb, vb = b[k]
        if not (restyling and k == edited):
            for (name, x), (_, y) in zip(fa, fb):
                if x != y:
                    problems.append('%s!%s の %s が %r -> %r' % (k[0], k[1], name, x, y))
        if va != vb and k != edited:
            problems.append('%s!%s の値が %r -> %r（編集対象外）' % (k[0], k[1], va, vb))

    if restyling:
        # The target must have changed FORMAT, and kept its value.
        if a.get(edited) and b.get(edited):
            if a[edited][0] == b[edited][0]:
                problems.append('編集対象 %s!%s の書式が変わっていない' % edited)
            if a[edited][1] != b[edited][1]:
                problems.append('書式だけの編集で %s!%s の値が変わった' % edited)
    elif a.get(edited) and b.get(edited) and a[edited][1] == b[edited][1]:
        problems.append('編集対象 %s!%s の値が変わっていない' % edited)

    for label, p in (('before', before_p), ('after', after_p)):
        for d in dangling_relationships(p):
            problems.append('%s のリンク切れ: %s' % (label, d))

    return problems


def main():
    if not os.path.isdir(PAIRS):
        print('error: %s がない。先に cargo test --lib xlsx_edit を実行する' % PAIRS)
        return 1

    present = set(os.listdir(PAIRS))
    stems = sorted(
        stem for stem in EDITED
        if {stem + '-before.xlsx', stem + '-after.xlsx'} <= present
    )
    if not stems:
        print('error: 比較する対がない。先に cargo test --lib xlsx_edit を実行する')
        return 1

    failed = 0
    for stem in stems:
        problems = check(stem)
        if problems:
            failed += 1
            print('FAIL %s' % stem)
            for p in problems[:30]:
                print('    %s' % p)
        else:
            print('ok   %s — 編集した1セル以外、書式も値も無変化' % stem)
    return 1 if failed else 0


if __name__ == '__main__':
    sys.exit(main())
