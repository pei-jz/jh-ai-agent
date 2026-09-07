# -*- coding: utf-8 -*-
"""Does a generated workbook come out readable when nobody said how?

That is the whole acceptance criterion for the presets: hand write_xlsx some
headers and some numbers, say NOTHING about layout, and the result should be
something you could print and hand to somebody.

Read by openpyxl rather than asserted in Rust for the same reason as
xlsx_fidelity.py — rust_xlsxwriter confirming its own output proves only that
it is self-consistent. Column widths in particular are a number this project
computes itself; the check that matters is what a different reader sees.

Run by `npm run test:xlsx`, after the Rust tests have written the samples into
src-tauri/target/fidelity/.
"""
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("skip: openpyxl is not installed (pip install openpyxl)")
    sys.exit(0)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAMPLES = os.path.join(ROOT, 'src-tauri', 'target', 'fidelity')

# Mirrors the sample in preset_tests: header row plus three data rows.
HEADERS = ['品目', '数量', '単価', '金額', '年', '備考']
LONGEST_ITEM = 'ステンレスねじ M4x10'   # 幅 = 半角7 + 全角7*2 = 21


def width_of(text):
    """Excel character units, counting CJK as two."""
    total = 0
    for ch in text:
        cp = ord(ch)
        wide = (0x1100 <= cp <= 0x115F or 0x2E80 <= cp <= 0x303E
                or 0x3041 <= cp <= 0x33FF or 0x3400 <= cp <= 0x4DBF
                or 0x4E00 <= cp <= 0x9FFF or 0xAC00 <= cp <= 0xD7A3
                or 0xF900 <= cp <= 0xFAFF or 0xFF00 <= cp <= 0xFF60)
        total += 2 if wide else 1
    return total


def check_default(ws, problems):
    # 1. 列幅 — 一番長い品目が入りきること。これが無いと ### や切れになる。
    a = ws.column_dimensions['A'].width
    need = width_of(LONGEST_ITEM)
    if not a or a < need:
        problems.append('A列の幅 %s が「%s」(幅%d) に足りない' % (a, LONGEST_ITEM, need))

    # 2. 表示形式 — 金額は桁区切り、年はそのまま。
    if ws['D2'].number_format != '#,##0':
        problems.append('金額列の表示形式が %r（#,##0 のはず）' % ws['D2'].number_format)
    if ws['E2'].number_format != 'General':
        problems.append('年の列に表示形式がついた: %r（2,026 になる）' % ws['E2'].number_format)

    # 3. 数値は右、文字は左。
    if ws['B2'].alignment.horizontal != 'right':
        problems.append('数量が右寄せでない: %r' % ws['B2'].alignment.horizontal)
    if ws['A2'].alignment.horizontal == 'right':
        problems.append('品目が右寄せになっている')

    # 4. 長文は折り返す。
    if not ws['F2'].alignment.wrap_text:
        problems.append('備考（長文）が折り返しになっていない')

    # 5. 見出しの体裁。
    h = ws['A1']
    if not h.font.bold:
        problems.append('見出しが太字でない')
    if not (h.fill and h.fill.fgColor and h.fill.fgColor.rgb and h.fill.fgColor.rgb.endswith('4472C4')):
        problems.append('見出しの塗りが %r' % (h.fill.fgColor.rgb if h.fill and h.fill.fgColor else None))
    if h.alignment.horizontal != 'center':
        problems.append('見出しが中央寄せでない')

    # 6. 罫線 — データ範囲に細線。
    if not (ws['B3'].border.left and ws['B3'].border.left.style):
        problems.append('データ範囲に罫線がない')

    # 7. 使える表になっているか。
    if ws.freeze_panes != 'A2':
        problems.append('見出し行が固定されていない: %r' % ws.freeze_panes)
    if not ws.auto_filter.ref:
        problems.append('オートフィルタがない')
    # openpyxl は絶対参照の形で返す。
    if str(ws.print_title_rows or '').replace('$', '') != '1:1':
        problems.append('印刷時に見出し行が繰り返されない: %r' % ws.print_title_rows)
    # 「1ページ幅に収める」は pageSetUpPr@fitToPage と fitToHeight=0 の組。
    # fitToWidth は既定が 1 なので、属性そのものは省略されうる。
    fit = ws.sheet_properties.pageSetUpPr
    if not (fit and fit.fitToPage):
        problems.append('fitToPage が立っていない')
    if ws.page_setup.fitToHeight != 0:
        problems.append('縦は成り行きのはず（潰すと読めない）: %r' % ws.page_setup.fitToHeight)
    if ws.page_setup.fitToWidth not in (None, 1):
        problems.append('横1ページになっていない: %r' % ws.page_setup.fitToWidth)


def check_plain(ws, problems):
    """plain は「他のプログラムが読む用」。飾りが付いていないこと。"""
    if ws.freeze_panes:
        problems.append('plain なのに枠が固定されている')
    if ws.auto_filter.ref:
        problems.append('plain なのにオートフィルタがある')
    if ws['A1'].fill and ws['A1'].fill.patternType:
        problems.append('plain なのに見出しが塗られている')
    if ws['B3'].border.left and ws['B3'].border.left.style:
        problems.append('plain なのに罫線がある')
    # 中身の判断（表示形式・幅）は plain でも効いていてよい。
    if ws['D2'].number_format != '#,##0':
        problems.append('plain でも桁区切りは効いてほしい: %r' % ws['D2'].number_format)


def check_report(ws, problems):
    if ws.page_setup.orientation != 'landscape':
        problems.append('report が横向きでない: %r' % ws.page_setup.orientation)
    band = ws['A3'].fill
    if not (band and band.patternType):
        problems.append('report に縞模様がない')


def check_override(ws, problems):
    """明示指定は preset に勝つ。"""
    if ws.freeze_panes:
        problems.append('freeze:0 を指定したのに固定されている: %r' % ws.freeze_panes)
    # Excel は列幅に既定フォントぶんの余白 (+0.7109375) を足して保存する。
    # 指定した 40 は 40.71 として書かれ、画面には 40 と出る。
    a = ws.column_dimensions['A'].width
    if a is None or abs(a - 40.0) > 1.0:
        problems.append('col_widths の A=40 が効いていない: %r' % a)


CHECKS = {
    'preset-default': check_default,
    'preset-plain': check_plain,
    'preset-report': check_report,
    'preset-override': check_override,
}


def main():
    if not os.path.isdir(SAMPLES):
        print('error: %s がない。先に cargo test --lib preset_tests を実行する' % SAMPLES)
        return 1

    failed = 0
    for stem, fn in CHECKS.items():
        path = os.path.join(SAMPLES, stem + '.xlsx')
        if not os.path.exists(path):
            print('FAIL %s — ファイルがない' % stem)
            failed += 1
            continue
        ws = load_workbook(path).worksheets[0]
        problems = []
        fn(ws, problems)
        if problems:
            failed += 1
            print('FAIL %s' % stem)
            for p in problems:
                print('    %s' % p)
        else:
            print('ok   %s' % stem)
    return 1 if failed else 0


if __name__ == '__main__':
    sys.exit(main())
