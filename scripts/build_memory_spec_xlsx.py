# -*- coding: utf-8 -*-
"""Memory feature specification — styled xlsx generator.

Builds docs/memory-feature-spec.xlsx with a designed layout:
  * 表紙   (cover): title band, field table, legend
  * 履歴   (history): version log table
  * 詳細仕様 (detailed spec): grouped section bands (merged cells),
    zebra stripes, borders, freeze panes, wrapped text, print setup.

Requires: openpyxl (pip install openpyxl)
"""
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "docs/memory-feature-spec.xlsx"

# --------------------------------------------------------------------------
# palette & styles
# --------------------------------------------------------------------------
NAVY = "1F3864"    # primary dark (headers / title)
INDIGO = "4472C4"  # section band (col A of the spec sheet)
LIGHT = "DCE6F1"   # light blue (cover labels / subtitle band)
STRIPE = "F7FAFD"  # zebra stripe
WHITE = "FFFFFF"
BORDER_COLOR = "8EAADB"
BODY_TEXT = "333333"

FONT = "Meiryo"

F_TITLE = Font(name=FONT, size=22, bold=True, color=WHITE)
F_SUB = Font(name=FONT, size=12, bold=True, color=NAVY)
F_HEAD = Font(name=FONT, size=11, bold=True, color=WHITE)
F_LABEL = Font(name=FONT, size=10, bold=True, color=NAVY)
F_BODY = Font(name=FONT, size=10, color=BODY_TEXT)
F_SECT = Font(name=FONT, size=10, bold=True, color=WHITE)
F_ITEM = Font(name=FONT, size=10, bold=True, color=NAVY)
F_FOOT = Font(name=FONT, size=9, italic=True, color="7F7F7F")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_INDIGO = PatternFill("solid", fgColor=INDIGO)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)
FILL_STRIPE = PatternFill("solid", fgColor=STRIPE)

_side = Side(style="thin", color=BORDER_COLOR)
BORDER_ALL = Border(left=_side, right=_side, top=_side, bottom=_side)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_LC = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_cell(ws, row, col, value, font, fill, alignment, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    if border:
        cell.border = BORDER_ALL
    return cell


def est_height(texts, widths):
    """Estimate a comfortable row height from wrapped CJK text lengths."""
    lines = 1
    for text, width in zip(texts, widths):
        if not text:
            continue
        # CJK glyphs are roughly half a column-width unit wide.
        per_line = max(4, int(width / 2))
        lines = max(lines, math.ceil(len(text) / per_line))
    return max(20, lines * 14 + 6)


wb = Workbook()

# --------------------------------------------------------------------------
# Sheet 1: 表紙 (cover)
# --------------------------------------------------------------------------
ws = wb.active
ws.title = "表紙"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 110

ws.merge_cells("A1:B1")
style_cell(ws, 1, 1, "メモリ機能 仕様書", F_TITLE, FILL_NAVY, AL_C)
ws.merge_cells("A2:B2")
style_cell(ws, 2, 1, "Memory Feature Specification ／ 実装コード準拠のリファレンス", F_SUB, FILL_LIGHT, AL_C)
ws.row_dimensions[1].height = 46
ws.row_dimensions[2].height = 24

style_cell(ws, 3, 1, "項目", F_HEAD, FILL_NAVY, AL_LC)
style_cell(ws, 3, 2, "内容", F_HEAD, FILL_NAVY, AL_LC)
ws.row_dimensions[3].height = 24

cover_rows = [
    ("文書名", "メモリ機能 仕様書（Memory Feature Specification）"),
    ("バージョン", "1.1"),
    ("作成日", "2026-08-14"),
    ("ステータス", "リファレンス（実装コード準拠）"),
    ("対象読者", "JHAI 本体の開発者、Settings → Memory 画面の利用者"),
    ("関連ドキュメント", "docs/design/agent-memory-layers.md（3層の設計）／docs/design/agent-memory-learning.md（学習設計）／docs/design/agent-memory-learning.plan.md（実行計画）"),
    ("概要", "エージェントの長期記憶を「誰が書き・いつ読まれるか」の違いで3層（Durable Facts / Experience / Episodes）に分け、加えて構造インデックス（Study）と概観ノート（Overview）をワークスペース単位で保持する。体験カードは実行トレースから機械的に導出され、LLMの自己申告が入らないのが特徴。"),
    ("主な実装モジュール", "src/modules/ai/memory/ 配下（CardStore.js / FactStore.js / FactExtraction.js / FailureSignature.js / MemoryScoring.js / SessionMetrics.js / TraceRecorder.js / ProjectOverview.js / StudyPass.js / CodeIndex.js / workspaceMemory.js）"),
    ("保存先", "<workspace>/.agent/ 配下（facts.json / memory.json / memory/cards.jsonl / memory/overview.md / trace/<sessionId>.jsonl / trace/metrics.jsonl）"),
    ("凡例", "semantic=確立済み意味記憶／episodic=試用中／norm=規範／observation=観測／worklog=作業ログ（保存されない）"),
    ("備考", "本仕様書は実装コード（JS）と設計ドキュメントを照合して作成。数値上限はコードの定数値を正とする。"),
]
row = 4
for label, value in cover_rows:
    style_cell(ws, row, 1, label, F_LABEL, FILL_LIGHT, AL_LC)
    style_cell(ws, row, 2, value, F_BODY, FILL_WHITE, AL_L)
    ws.row_dimensions[row].height = est_height([value], [110])
    row += 1

ws.print_area = f"A1:B{row - 1}"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

# --------------------------------------------------------------------------
# Sheet 2: 履歴 (history)
# --------------------------------------------------------------------------
ws2 = wb.create_sheet("履歴")
ws2.sheet_properties.tabColor = "2F5597"
ws2.sheet_view.showGridLines = False
for col, width in zip("ABCD", (12, 14, 85, 18)):
    ws2.column_dimensions[col].width = width

headers2 = ["バージョン", "日付", "変更内容", "作成者"]
for c, h in enumerate(headers2, start=1):
    style_cell(ws2, 1, c, h, F_HEAD, FILL_NAVY, AL_LC)
ws2.row_dimensions[1].height = 24

history_rows = [
    ("1.1", "2026-08-14", "デザイン改訂。背景色・フォント・罫線・セクション結合・ゼブラストライプを適用し、設計書としての可読性を向上（内容は1.0から不変）", "JH AI Agent"),
    ("1.0", "2026-08-14", "初版。3層記憶＋構造インデックス＋概観ノートの仕様を実装コードからドキュメント化", "JH AI Agent"),
]
row = 2
for version, date, note, author in history_rows:
    style_cell(ws2, row, 1, version, F_ITEM, FILL_WHITE, AL_LC)
    style_cell(ws2, row, 2, date, F_BODY, FILL_WHITE, AL_LC)
    style_cell(ws2, row, 3, note, F_BODY, FILL_WHITE, AL_L)
    style_cell(ws2, row, 4, author, F_BODY, FILL_WHITE, AL_LC)
    ws2.row_dimensions[row].height = est_height([note], [85])
    row += 1

ws2.freeze_panes = "A2"
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr.fitToPage = True

# --------------------------------------------------------------------------
# Sheet 3: 詳細仕様 (detailed spec)
# --------------------------------------------------------------------------
ws3 = wb.create_sheet("詳細仕様")
ws3.sheet_properties.tabColor = INDIGO
ws3.sheet_view.showGridLines = False
for col, width in zip("ABC", (24, 30, 120)):
    ws3.column_dimensions[col].width = width

spec = [
    # (大分類, 項目, 仕様詳細)
    ("1. 全体像", "設計思想", "記憶は「情報の種類」ではなく「誰が書き、いつ読まれるか」で層分けする。Facts と Episodes は LLM がセッション終了時に自己申告、Experience は実行トレースから機械的に導出（LLM 不使用）。"),
    ("1. 全体像", "3層の比較", "Durable Facts=プロジェクトの取扱説明書（何か）／Experience=どうやるか・どこにあるか・何で転んだか／Episodes=いつ何をしたか（作業日誌）"),
    ("1. 全体像", "次タスクでの注入", "Facts: システムプロンプトに上位5件。Episodes: 関連度0.08以上の上位3件を時系列で。Experience: 冒頭ブリーフ上位3枚＋ツール呼び出し時に1枚"),
    ("1. 全体像", "保持上限", "Facts: 100件／Experience: 300枚／Episodes: 直近20件（FIFO）"),
    ("1. 全体像", "消え方", "Facts: hits×層重み×90日半減期で減衰。Experience: costSteps×hits×confidence×90日半減期。Episodes: 古いものから順に削除"),
    ("2. 保存ファイル", "ディレクトリ", "<workspace>/.agent/ 配下に全て格納。パスは workspaceMemory.memoryPaths() が一元管理"),
    ("2. 保存ファイル", "facts.json", "JSON配列、全体書き換え。Durable Facts の保存先。書式: [{\"fact\",\"kind\",\"hits\",\"type\",\"category\",\"scope\":\"workspace\",\"evidence\"}]"),
    ("2. 保存ファイル", "memory.json", "JSON配列、全体書き換え。Episodes の保存先。1行=1セッション要約（topic/actions/outcome/keyFiles/category/summary/facts）"),
    ("2. 保存ファイル", "memory/cards.jsonl", "JSON Lines（1行1カード）。Experience の保存先。末尾改行必須（追記用）。壊れた行は読み飛ばす"),
    ("2. 保存ファイル", "memory/overview.md", "Markdown（人が手で読んで修正するため）。先頭フロントマターに generated / head / conventions を保持"),
    ("2. 保存ファイル", "trace/<sessionId>.jsonl", "1セッションのツール呼び出し記録（Experience の原料）。20件ごとにバッファフラッシュ"),
    ("2. 保存ファイル", "trace/metrics.jsonl", "1セッション1行の計測データ（A/B比較の原料）。読み取り→追記方式で並行タスクに耐える"),
    ("3. Durable Facts", "記銘（書き込み）", "セッション終了時に LLM が JSON で要約。facts 配列は最大3件、各300字。FactExtraction.buildSummaryPrompt() がプロンプトを生成"),
    ("3. Durable Facts", "fact の kind", "norm=プロジェクト規範（1回で semantic に昇格）／observation=観測（同一カテゴリ内3回で semantic）／worklog=作業ログ（保存しない）。ユーザーによる訂正は最優先で norm として記録"),
    ("3. Durable Facts", "昇格ルール", "applyPromotion(): norm は即 semantic（confidence=0.5+0.15×(hits-1)、上限0.9）。observation は hits>=3 で semantic（confidence=0.7）、未満は episodic（0.4）。type 未指定は semantic 扱い（後方互換）"),
    ("3. Durable Facts", "マージ", "mergeFacts(): 正規化テキスト完全一致 or Jaccard>=0.7 かつ同一カテゴリで重複判定。8文字未満は破棄。重複時は hits+1、kind は norm へ一方通行で昇格"),
    ("3. Durable Facts", "選出", "selectRelevantFacts(): relevanceScore>=0（任意閾値）の上位5件。関連度優先→confidence→timestamp でソート。selectNormFacts(): 規範は関連度フィルタなしで別予算3件（キーワードに現れなくても常に注入）"),
    ("3. Durable Facts", "保持と刈り込み", "retentionScore()=hits×層重み(semantic/procedural=1, episodic=0.6)×0.5^(経過日/90)。pruneFacts() が100件超をスコア順に刈る"),
    ("3. Durable Facts", "統合", "applyConsolidation(): LLM の統合プラン {remove:[], merge:[{into,from,text}]} を適用。70%超の一括削除は安全弁で拒否"),
    ("4. Experience（カード）", "カード種別", "lesson=失敗とその対処（costSteps 実測済み）／insight=成功した手順・場所。insight の下位種別: recovery（成功手順）と locator（検索→使用実績のある場所）"),
    ("4. Experience（カード）", "生成", "mintCards(): TraceRecorder.summarizeFailures() の行と生イベントから機械的に生成。LLM は一切介在しない。denied / permission_denied はユーザー拒否なので学習しない"),
    ("4. Experience（カード）", "レシピ（fix）", "recoveryRecipe(): 失敗後、同じツールが同じターゲットで成功するまでの連続成功ツール列。空なら fix=null（未検証）。2ステップ以上で insight/recovery カードも生成"),
    ("4. Experience（カード）", "locator", "isDurableQuery() で正規表現・glob 等の一回使いクエリを除外（名前のみ残す）。検索後3ステップ以内に read/write 系で結果を使用した場合のみ記録（使われなかった検索は残さない）"),
    ("4. Experience（カード）", "署名（signature）", "FailureSignature.signatureOf(): 「ツール名|エラー種別|拡張子」の可読文字列（例: write_file|edit_mismatch|.svelte）。ハッシュ化はしない。カードIDのみ FNV-1a fingerprint() で8桁hex"),
    ("4. Experience（カード）", "マージ", "mergeCards(): cardKey 一致で hits+1、last_recurrence 更新。injected 中に再発した場合のみ recurrences_after_hit+1（提示していない回の再発はカードに課さない）"),
    ("4. Experience（カード）", "並行調整", "reconcile(): ディスク上の他セッション分と統合。hits は max を採用（mergeCards と違いカウントしない）。disabled はどちらかが OFF なら OFF"),
    ("4. Experience（カード）", "スコア", "cardScore()=max(1,costSteps)×hits×confidence×0.5^(経過日/90)。disabled / stale は0"),
    ("4. Experience（カード）", "想起（ツール時）", "selectForTool(): 実行予定ツール名と拡張子が一致するカードのうち最高スコア1枚。カード側に拡張子がない場合は全ツール対象、両方ある場合は一致必須（.svelte の教訓を .rs 編集で流用しない）"),
    ("4. Experience（カード）", "想起（冒頭）", "selectBriefBudgeted(): 種類別予算 BRIEF_BUDGET={insight:2, lesson:1}。クエリ指定時は関連度優先、なければスコア順。costSteps 主導だと痛い些事がルールを追い出すため分離予算"),
    ("4. Experience（カード）", "注入テキスト", "renderCard(): 最大240字。lesson は「[Memory — cost N steps here before] ... What worked: ... Do that first.」。locator は「\"X\" was found in Y. Look there first.」。否定形でなく実行指示形で書く"),
    ("4. Experience（カード）", "再発率", "recurrenceRate()=recurrences_after_hit÷shown（提示済みのみ分母）。0=効いている／1=効いていない／null=未提示で判定不能"),
    ("4. Experience（カード）", "保存", "CardStore クラス: load() で injected を除去（セッション単位フラグ）。save() で読み直し→reconcile→300枚超をスコア刈り→JSONL書き込み。maxPerRun=6（1実行で提示するカード上限）"),
    ("5. Episodes", "記銘", "セッション終了時に LLM が JSON で要約。topic=40字以内／actions=最大3件／outcome=success|partial|error／keyFiles=最大3件／summary=120字以内"),
    ("5. Episodes", "保存", "memory.json に直近20件を保持。超えた分は古いものから FIFO で破棄"),
    ("5. Episodes", "想起", "クエリとの関連度（relevanceScore）が0.08未満なら1件も入らない。関連する上位3件を時系列に並べ直してシステムプロンプトへ"),
    ("5. Episodes", "今後の方針", "Step 4 でプロンプト注入をやめ、知識グラフのエッジ生成の原料に降格予定（設計ドキュメント記載。実装は未着手）"),
    ("6. 構造インデックス（Study）", "目的", "エージェントが歩いた場所しか覚えない Experience の欠点を補い、未実行ワークスペースの構造を事前学習する"),
    ("6. 構造インデックス（Study）", "対象", "STUDY_GLOB=**/*.{js,jsx,mjs,cjs,ts,tsx,rs,py,java}、ファイル上限1000、1ファイル12シンボル。xlsx/xlsm は数式グラフ用に上限60"),
    ("6. 構造インデックス（Study）", "公平選択", "fairShare(): ディレクトリバケツをラウンドロビンで均等配分。大規模モノレポで1ディレクトリが独占しないようにする"),
    ("6. 構造インデックス（Study）", "出力", "シンボル定義（file:line）＋import エッジ＋ワークブック数式エッジ。Rust 側の index.db（SQLite）にキャッシュ、.agent/memory/ 配下"),
    ("6. 構造インデックス（Study）", "LLM不使用", "記録するのは解析で確定できる事実のみ（シンボル宣言位置・ファイル数）。推論は概観ノート生成に分離"),
    ("7. 概観ノート（Overview）", "目的", "「このプロジェクトは何か」を数百トークンで常時注入。粗くて完全な全体像レイヤ（GIST層）"),
    ("7. 概観ノート（Overview）", "上限", "OVERVIEW_MAX_CHARS=1600字。エリア分割 AREA_LIMIT=14、エリアあたりエクスポート名6個"),
    ("7. 概観ノート（Overview）", "フロントマター", "<!-- generated: ISO日時 -->／<!-- head: HEADコミット -->／<!-- conventions: JSON -->。conventions は計測済みレイヤでLLMなしで更新可能"),
    ("7. 概観ノート（Overview）", "命名規則検出", "detectConventions(): 既存5種（Java/MyBatis形状＋テスト命名）＋一般化（PascalCase/camelCase/snake_case/kebab-case、接頭辞タクソノミー、共有中間語、生成ディレクトリ除外、慣用ディレクトリ役割）"),
    ("7. 概観ノート（Overview）", "カバレッジ", "detectConventionsFull(): rules の coverage を計算。ソースファイル（アセット・テスト・生成物を除外後）のうち規則で説明できる割合。recipe（検索レシピ）も付与"),
    ("7. 概観ノート（Overview）", "生成タイミング", "Settings → Memory → Study 実行後。LLM に構造ダイジェストだけを見せて要約させる（生ソースは見せない）"),
    ("8. スコアリング", "textUnits", "ラテン単語（3文字以上）＋CJK 文字バイグラムの集合に分割。日本語で Jaccard が常に0になる旧実装の欠陥を修正"),
    ("8. スコアリング", "relevanceScore", "クエリ単位のうちエントリ本文に含まれる割合（0〜1）。空クエリは0.5。閾値 MEMORY_MIN_RELEVANCE=0.08"),
    ("8. スコアリング", "sanitizeXmlTags", "システムプロンプトの構造化セクションを汚染しないよう、注入テキストのタグを [tag] 形式に置換"),
    ("9. トレースと計測", "TraceRecorder", "全ツール呼び出しをイベント記録（成功も失敗も）。argShape・target（redact済み）・ms・検索クエリを保持。toEvent() は純関数で単体テスト対象"),
    ("9. トレースと計測", "エラー正規化", "normalizeError(): ①redact（秘密/PIIマスク）→②normalizeMessage（パス・ハッシュ・タイムスタンプ・位置数字を <path>/<hash>/<ts>/<n> に置換）→③errorKind（13種の正規表現テーブル、最初に一致したものを採用）"),
    ("9. トレースと計測", "失敗要約", "summarizeFailures(): (signature,target) 単位に集約。costSteps=初回失敗イテレーションから同ツール同ターゲットの成功イテレーションまで。未解決はセッション末尾まで課金し unresolved=true"),
    ("9. トレースと計測", "SessionMetrics", "セッション毎に explorationCost（初回編集前のツール呼び出し数）・reReads（同一ファイル再読回数）・followThrough（提示カードの手順を部分列一致で追従したか）を計測"),
    ("9. トレースと計測", "A/B比較", "compareArms(): recall=on 群と off 群で平均比較。delta は on−off（マイナスが改善）。followThrough.lift=on率−off率。両群にデータが無い限り delta は null（片群だけの結果は出さない）"),
    ("9. トレースと計測", "必要サンプル数", "runsNeeded(): 80%検出力・α=0.05 で n=15.7σ²/Δ²。5件未満は null（推測で出さない）"),
    ("10. 思い出し設定", "memory_recall 設定", "Settings → General → Agent Safety Limits → Memory Recall。on=常に想起（既定）／auto=約10%を無作為に対照群に／off=想起しない（学習は継続）"),
    ("10. 思い出し設定", "shadow 選択", "対照群はカードを選択するが注入しない（shadow=true）。shown にも injected にも数えない（再発率の分母を水増ししない）"),
    ("11. UI（MemoryTab）", "構成", "ワークスペース選択（入力＋datalist＋参照ボタン）→ Load ボタンで読み込み。Study ボタンで構造学習開始。セクションは折りたたみ可能（初期は全て開く）"),
    ("11. UI（MemoryTab）", "Facts 表示", "一覧表示＋行内インライン編集（window.prompt ではなく）。種別バッジ: semantic/episodic、norm/observation。全削除ボタンあり"),
    ("11. UI（MemoryTab）", "Cards 表示", "lesson/insight/locator のバッジと要約行。OFF トグル（disabled=true でスコア0、注入されなくなる。削除でなく無効化＝追跡可能性を保つ）、個別削除・全削除"),
    ("11. UI（MemoryTab）", "Episodes 表示", "outcome アイコン（✅/❌/⚠️）付き一覧、個別削除・全削除"),
    ("11. UI（MemoryTab）", "A/B 進捗表示", "abStats から必要サンプル数に対する現在の走行数をバー表示。十分なデータが無いうちは「測定中」と正直に表示"),
    ("11. UI（MemoryTab）", "Overview 表示", "概観ノートのテキストをインライン編集可能（編集→保存で writeOverview）。head・generatedAt のスタンプ情報も表示"),
    ("11. UI（MemoryTab）", "Index 統計", "files/symbols/edges 数とカバレッジを表示（未学習領域は「推測で答える領域」として明示）"),
    ("12. パスガード", "allowMemoryDir", "書き込み前に <workspace>/.agent を set_allowed_roots で登録（読み取りは非ゲート）。古いバックエンドでは失敗しても握りつぶす（書き込み不可の環境で保存を止めない）"),
    ("12. パスガード", "読み書きAPI", "Rust コマンド（read_file / write_file / create_dir）経由。フロントは invoke を注入し、テストは Tauri 不要"),
    ("13. 並行性", "同一ファイル書き込み", "カードは save() 前に再読取→reconcile でマージ（サブエージェントや並行タスクの学習を消さない）。metrics は追記、facts/episodes は全体書換だが実害を最小化"),
    ("13. 並行性", "1実行1回原則", "同じカードは1実行につき1度だけ提示（injected セット）。maxPerRun=6 で過剰な注入密度を防止"),
    ("14. テスト", "テスト方針", "各モジュールの純関数を単体テストで固定（*.test.js）。特に FailureSignature の正規化テーブル、CardStore のスコア/マージ、FactStore の昇格/統合、MemoryScoring のトークン化、SessionMetrics の A/B ロジック"),
    ("14. テスト", "テスト対象ファイル", "src/modules/ai/memory/__tests__/ 配下。MemoryTab は src/dashboard/svelte/config/__tests__/tabs.test.js で検証"),
]

# header row
for c, h in enumerate(["大分類", "項目", "仕様詳細"], start=1):
    style_cell(ws3, 1, c, h, F_HEAD, FILL_NAVY, AL_LC)
ws3.row_dimensions[1].height = 26

r = 2
group_idx = 0
for i, (cat, item, detail) in enumerate(spec):
    # section band: merge the category column across consecutive rows
    if i == 0 or spec[i - 1][0] != cat:
        start = r
    if i == len(spec) - 1 or spec[i + 1][0] != cat:
        end = r
        ws3.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        for rr in range(start, end + 1):
            ws3.cell(row=rr, column=1).fill = FILL_INDIGO
            ws3.cell(row=rr, column=1).border = BORDER_ALL
        top = ws3.cell(row=start, column=1)
        top.value = cat
        top.font = F_SECT
        top.alignment = AL_LC
        group_idx += 1

    zebra = FILL_WHITE if (r % 2 == 0) else FILL_STRIPE
    style_cell(ws3, r, 2, item, F_ITEM, zebra, AL_LC)
    style_cell(ws3, r, 3, detail, F_BODY, zebra, AL_L)
    ws3.row_dimensions[r].height = est_height([item, detail], [30, 120])
    r += 1

# footer note
last = r
ws3.merge_cells(start_row=last, start_column=1, end_row=last, end_column=3)
style_cell(ws3, last, 1, "出典: 実装コード（src/modules/ai/memory/）および設計ドキュメント（docs/design/agent-memory-*.md）", F_FOOT, FILL_LIGHT, AL_LC)
ws3.row_dimensions[last].height = 20

ws3.freeze_panes = "A2"
ws3.page_setup.orientation = "landscape"
ws3.page_setup.fitToWidth = 1
ws3.page_setup.fitToHeight = 0
ws3.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUT)
print(f"written: {OUT}")
print(f"  cover rows    : {len(cover_rows) + 3}")
print(f"  history rows  : {len(history_rows) + 1}")
print(f"  spec rows     : {len(spec) + 2} (data {len(spec)} + header + footer)")
