# -*- coding: utf-8 -*-
"""Does an appended row really look like the row above it?

The acceptance criterion for "既存書式を踏襲した修正": add one line to a 明細
and the new line should be indistinguishable from the last one — same ruling,
same number format, same alignment, same height — with only the values
different.

Checked by openpyxl rather than in Rust for the usual reason: the Rust side can
only confirm that it copied the style INDEX it meant to copy. Whether that index
actually means "thin border, #,##0円, right-aligned" is a question for a reader
that has no stake in the answer.

Run by `npm run test:xlsx`, after the Rust tests have written the samples.
"""
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("skip: openpyxl is not installed (pip install openpyxl)")
    sys.exit(0)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAMPLES = os.path.join(ROOT, 'src-tauri', 'target', 'fidelity')

TEMPLATE_ROW = 4      # the last row of the fixture
NEW_ROW = 5           # the row the append adds
COLUMNS = 'ABCD'


def look(c):
    """Everything about a cell except what it says."""
    b, a = c.border, c.alignment
    return {
        'border': tuple((getattr(b, s).style if getattr(b, s) else None)
                        for s in ('left', 'right', 'top', 'bottom')),
        'numfmt': c.number_format,
        'halign': a.horizontal,
        'valign': a.vertical,
        'wrap': a.wrap_text,
        'bold': c.font.bold,
        'font': c.font.name,
        'fill': (c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None),
    }


def main():
    before_p = os.path.join(SAMPLES, 'ledger-before.xlsx')
    after_p = os.path.join(SAMPLES, 'ledger-appended.xlsx')
    if not os.path.exists(after_p):
        print('error: %s がない。先に cargo test --lib xlsx_edit を実行する' % after_p)
        return 1

    before = load_workbook(before_p).worksheets[0]
    after = load_workbook(after_p).worksheets[0]
    problems = []

    # 1. 追加した行が、上の行と同じ見た目か。
    for col in COLUMNS:
        want = look(after['%s%d' % (col, TEMPLATE_ROW)])
        got = look(after['%s%d' % (col, NEW_ROW)])
        for k in want:
            if want[k] != got[k]:
                problems.append('%s%d の %s が上の行と違う: %r ≠ %r'
                                % (col, NEW_ROW, k, got[k], want[k]))

    # 2. 罫線が行の端まで続いているか（値を入れなかった列も含めて）。
    for col in COLUMNS:
        c = after['%s%d' % (col, NEW_ROW)]
        if not (c.border.left and c.border.left.style):
            problems.append('%s%d に罫線がない（表の途中で線が切れる）' % (col, NEW_ROW))

    # 3. 行の高さも引き継ぐ。
    h_before = before.row_dimensions[TEMPLATE_ROW].height
    h_after = after.row_dimensions[NEW_ROW].height
    if h_before is not None and h_after != h_before:
        problems.append('行の高さが引き継がれていない: %r ≠ %r' % (h_after, h_before))

    # 4. 値は入っていること。数式は数式のまま。
    if after['A%d' % NEW_ROW].value != 'ワッシャー':
        problems.append('値が入っていない: %r' % after['A%d' % NEW_ROW].value)
    d = after['D%d' % NEW_ROW].value
    if not (isinstance(d, str) and d.startswith('=')):
        problems.append('数式が数式として入っていない: %r' % d)

    # 5. 既存の行には触っていないこと。
    for row in range(1, TEMPLATE_ROW + 1):
        for col in COLUMNS:
            a = before['%s%d' % (col, row)]
            b = after['%s%d' % (col, row)]
            if a.value != b.value or look(a) != look(b):
                problems.append('既存の %s%d が変わった' % (col, row))

    if problems:
        print('FAIL ledger-appended')
        for p in problems[:20]:
            print('    %s' % p)
        return 1
    print('ok   ledger-appended — 追加行は上の行と同じ書式、既存行は無変化')
    return 0


if __name__ == '__main__':
    sys.exit(main())
