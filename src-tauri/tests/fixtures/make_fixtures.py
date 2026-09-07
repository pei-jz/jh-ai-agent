# -*- coding: utf-8 -*-
"""Rebuild the .xlsx test fixtures.

The point of these files is that WE did not write them. An editor verified
against workbooks produced by its own writer only proves the writer is
self-consistent; the failures this suite exists for came from real files —
charts, pivot caches, shared formulas, merged ranges — produced by something
else. openpyxl stands in for "a real spreadsheet tool" here.

    pip install openpyxl
    python src-tauri/tests/fixtures/make_fixtures.py

The outputs are committed. Regenerate only when a fixture needs to cover
something new, and re-run the fidelity tests afterwards.
"""
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def styled():
    """A ruled, formatted table — the shape most business sheets actually are."""
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    for col, name in zip("ABC", ["品目", "数量", "金額"]):
        c = ws[col + "1"]
        c.value = name
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.border = BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, (n, q) in enumerate([("ねじ", 10), ("板金", 4)], start=2):
        ws[f"A{i}"] = n
        ws[f"B{i}"] = q
        ws[f"C{i}"] = f"=B{i}*25"
        ws[f"C{i}"].number_format = '#,##0"円"'
        for col in "ABC":
            ws[col + str(i)].border = BOX
    ws.merge_cells("A5:C5")
    ws["A5"] = "合計欄"
    ws.freeze_panes = "A2"
    wb.save(os.path.join(HERE, "styled.xlsx"))


def rich():
    """Everything an edit used to destroy, in one file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"

    for col, name in zip("ABCD", ["品目", "数量", "単価", "金額"]):
        c = ws[col + "1"]
        c.value = name
        c.font = Font(bold=True, color="FFFFFF", name="Meiryo")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX

    for i, (name, qty, price) in enumerate(
            [("ねじ", 10, 25), ("板金", 4, 1800), ("塗料", 2, 3200)], start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = qty
        ws[f"C{i}"] = price
        ws[f"D{i}"] = f"=B{i}*C{i}"
        ws[f"C{i}"].number_format = '#,##0"円"'
        ws[f"D{i}"].number_format = '#,##0"円"'
    ws["D5"] = "=SUM(D2:D4)"
    ws["A7"] = "納期"
    ws["B7"] = "2026-09-30"
    ws["B7"].number_format = "yyyy/mm/dd"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:D4"
    ws.merge_cells("A9:D9")
    ws["A9"] = "備考欄"
    ws["A11"] = "リンク"
    ws["A11"].hyperlink = "https://example.com"
    ws["B2"].comment = Comment("在庫僅少", "検査")

    ws.conditional_formatting.add(
        "B2:B4",
        CellIsRule(operator="lessThan", formula=["5"],
                   fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add(
        "D2:D4",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="63BE7B"))

    dv = DataValidation(type="list", formula1='"可,不可"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("E2:E4")

    ws2 = wb.create_sheet("集計")
    ws2["A1"] = "項目"
    ws2["B1"] = "値"
    for i, (k, v) in enumerate([("A", 30), ("B", 50), ("C", 20)], start=2):
        ws2[f"A{i}"] = k
        ws2[f"B{i}"] = v
    ch = BarChart()
    ch.title = "内訳"
    ch.add_data(Reference(ws2, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    ch.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=4))
    ws2.add_chart(ch, "D2")

    ws3 = wb.create_sheet("表")
    ws3.append(["ID", "名前"])
    ws3.append([1, "あ"])
    ws3.append([2, "い"])
    t = Table(displayName="明細表", ref="A1:B3")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws3.add_table(t)

    wb.save(os.path.join(HERE, "rich.xlsx"))


def ledger():
    """A plain 明細: header, ruled data rows, nothing after them.

    The fixture for appending. styled.xlsx ends in a merged 合計欄, which is a
    fine thing to test against but a poor template for "one more line".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "明細"
    for col, name in zip("ABCD", ["品目", "数量", "単価", "金額"]):
        c = ws[col + "1"]
        c.value = name
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.border = BOX
        c.alignment = Alignment(horizontal="center")
    for i, (n, q, p) in enumerate([("ねじ", 10, 25), ("板金", 4, 1800), ("塗料", 2, 3200)], start=2):
        ws[f"A{i}"] = n
        ws[f"B{i}"] = q
        ws[f"C{i}"] = p
        ws[f"D{i}"] = f"=B{i}*C{i}"
        for col in "ABCD":
            ws[col + str(i)].border = BOX
        ws[f"C{i}"].number_format = '#,##0"円"'
        ws[f"D{i}"].number_format = '#,##0"円"'
        ws[f"B{i}"].alignment = Alignment(horizontal="right")
        ws.row_dimensions[i].height = 22
    wb.save(os.path.join(HERE, "ledger.xlsx"))


if __name__ == "__main__":
    styled()
    rich()
    ledger()
    for f in ("styled.xlsx", "rich.xlsx", "ledger.xlsx"):
        print("%-14s %6d bytes" % (f, os.path.getsize(os.path.join(HERE, f))))
